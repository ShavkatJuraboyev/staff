import openpyxl, requests, json # Excel fayllarini o'qish uchun kutubxona
from django.shortcuts import render, redirect,get_object_or_404 # render va redirect kutubxonasini import qilamiz
from employee.models import Employee, Role, UserRole, Departments # Talabalar modelini import qilamiz
from django.contrib import messages # Xabarlar uchun
from django.db.models import Q # Q kutubxonasini import qilamiz
from django.core.paginator import Paginator # Paginator kutubxonasini import qilamiz
from django.http import HttpResponseNotFound, HttpResponse, JsonResponse # 404 xatolikni chiqarish uchun
from django.db import IntegrityError
from django.contrib.auth.models import User
from employee.decorators import role_required # Ruxsatlarni tekshirish uchun
from django.contrib.auth.decorators import login_required # Login bo'lishni tekshirish uchun
from django.contrib.auth import login, logout, authenticate # Login, logout va authenticate uchun
from employee.forms import LoginForm # Login formasi
from itertools import groupby
from operator import attrgetter

def login_decorator(func):
    return login_required(func, login_url='user_login')

def user_login(request):
    if request.POST:
        forms = LoginForm(request.POST)
        if forms.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, password=password, username=username)
            if user is not None:
                login(request, user)
                messages.success(request, f'{request.user.username} tizimga kirdiz!')
                return redirect('all_employees')
            else:
                messages.warning(request, 'Bunday foydalanuvchi mavjud emas!')
                return redirect('user_login')
        else:
            messages.warning(request, 'Username yoki parol xato!')
            return redirect('user_login')
    else:
        forms = LoginForm()
    return render(request, 'admin/login.html')

def get_employment_info(hemis_id): # Hemis ID bo'yicha ish joyi ma'lumotlarini olish
    """Hemis id kiritilishi lozim"""
    url = f"https://student.samtuit.uz/rest/v1/data/employee-list?type=all&search={hemis_id}" # API URL
    token = "Y-R36P1BY-eLfuCwQbcbAlvt9GAMk-WP" # Token
    headers = {"Authorization": "Bearer " + token} # Tokenni headerga qo'shish
    response = requests.get(url, headers=headers) # API ga so'rov yuborish
    if response.status_code == 200: # Agar so'rov muvaffaqiyatli bo'lsa
        return response.json() # JSON ma'lumotlarni qaytarish
    return None # Aks holda None qaytarish

@login_decorator
def logOut(request):
    logout(request)
    return redirect("user_login")
 
@login_decorator
@role_required("add_user")
def upload_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Excel qatoridagi ma'lumotlar
                employee_data = {
                    'employee_id': row[0],
                    'citizenship': row[1],
                    'passport': row[2],
                    'personal_number': row[3],
                    'first_name': row[6],
                    'last_name': row[7],
                    'sur_name': row[8],
                    'academic_degree': row[12],
                    'academic_title': row[13],
                    'bithday': row[20],
                    'gender': row[22],
                    'email': row[23],
                    'phone': row[24],
                    'permanent_registration': row[25]
                }

                department_data = {
                    'faculty': row[4],
                    'department': row[5],
                    'labor_form': row[9],
                    'stavka': row[10],
                    'position': row[11],
                    'expertise': row[14],
                    'start_position': row[15],
                    'contract_number': row[16],
                    'contract_date': row[17],
                    'order_number': row[18],
                    'order_date': row[19],
                }

                # Employee ma'lumotlarini saqlash yoki yangilash
                employee, created = Employee.objects.update_or_create(
                    employee_id=employee_data['employee_id'],
                    defaults=employee_data,
                )

                # Departments ma'lumotlarini saqlash yoki yangilash
                if employee:
                    Departments.objects.update_or_create(
                        employees=employee,
                        faculty=department_data['faculty'],
                        department=department_data['department'],
                        defaults=department_data,
                    )

            messages.success(request, "Ma'lumotlar muvaffaqiyatli yuklandi!")
        except IntegrityError as e:
            messages.error(request, f"Xato yuz berdi: {str(e)}")
        except Exception as e:
            messages.error(request, f"Xatolik: {str(e)}")

        return redirect("upload_excel")

    return render(request, "admin/upload_excel.html")

@login_decorator
@role_required("download_user")
def export_employees_to_excel(request):
    query = request.GET.get('q', '')  # Qidiruv uchun so'rov
    katta_yosh = request.GET.get('katta', None)  # Katta yosh filtri
    kichik_yosh = request.GET.get('kichik', None)  # Kichik yosh filtri
    gender = request.GET.get('gender', None)  # Jinsi
    ilmiy_daraja = request.GET.get('ilmiy_daraja', None)  # Ilmiy daraja
    ilmiy_unvon = request.GET.get('ilmiy_unvon', None)  # Ilmiy unvon
    kafedra = request.GET.get('kafedra', None)  # kafedra
    lavozim = request.GET.get('lavozim', None)  # kafedra
    tugilgan_joyi = request.GET.get('placeofbirtht', None)  # viloyati
    city = request.GET.get('city', None)  # Doimiy yashash joyi

    # Xodimlar ma'lumotlarini olish
    employees = Employee.objects.filter(xodim__isnull=False).distinct().prefetch_related("xodim").order_by('-id')

    # Filtrlash
    if query:
        employees = employees.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(sur_name__icontains=query) |
            Q(employee_id__icontains=query) |
            Q(phone__icontains=query) |
            Q(age__icontains=query) 
        )
    if katta_yosh:
        employees = employees.filter(age__gte=int(katta_yosh))  # Yoshdan katta
    if kichik_yosh:
        employees = employees.filter(age__lte=int(kichik_yosh))  # Yoshdan kichik
    if gender:
        employees = employees.filter(gender=gender)  # Jinsi bo'yicha filtr
    if ilmiy_daraja:
        employees = employees.filter(academic_degree__icontains=ilmiy_daraja)  # Ilmiy daraja bo'yicha filtr
    if ilmiy_unvon:
        employees = employees.filter(academic_title__icontains=ilmiy_unvon)  # Ilmiy unvon bo'yicha filtr
    if kafedra:
        employees = employees.filter(xodim__department__icontains=kafedra)  # Kafedra bo'yicha filtr
    if lavozim:
        employees = employees.filter(xodim__labor_form__icontains=lavozim)  # lavozim bo'yicha filtr
    if tugilgan_joyi:
        employees = employees.filter(place_of_birth__icontains=tugilgan_joyi)  # Tug'ilgan joyi filtr
    if city:
        employees = employees.filter(city__icontains=city) # Doimiy yashash joyi filtr

    # Excel faylni yaratamiz
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    # Sarlavha qatorini qo'shamiz
    headers = [
        "Xodim ID", "Fuqaroligi", "Passport seriyasi", "Shaxsiy raqam (JSHER)", "Fakulteti",
        "Kafedra/Bo'lim", "Ismi", "Familiyasi", "Otasining ismi", "Ish shakli", "Stavka",
        "Lavozimi", "Ilmiy darajasi", "Ilmiy unvon", "Mutaxassisligi", "Ishga kirgan sana",
        "Shartnoma raqami", "Shartnoma sanasi", "Buyruq raqami", "Buyruq sanasi", 
        "Tug'ilgan kun", "Yoshi", "Jinsi", "Email", "Telefon raqami", "Doimiy yashash joyi",
        "Tashkiloti nomi", "Millati", "Tug'ilgan joyi", "Tuman yoki Shaxar",
        "Tugatgan yili", "Tugatgan oliy ta'lim nomi"
    ]
    sheet.append(headers)

    # Ma'lumotlarni qo'shamiz
    for employee in employees:
        departments = Departments.objects.filter(employees=employee)

        def get_unique_values(field_name):
            values = set(departments.values_list(field_name, flat=True))
            values.discard(None)
            return ", ".join(values) if values else ""

        sheet.append([  # Employee ma'lumotlarini qo'shamiz
            employee.employee_id, employee.citizenship, employee.passport, employee.personal_number,
            get_unique_values("faculty"), get_unique_values("department"),
            employee.first_name, employee.last_name, employee.sur_name,
            get_unique_values("labor_form"), get_unique_values("stavka"),
            get_unique_values("position"), employee.academic_degree, employee.academic_title,
            get_unique_values("expertise"), get_unique_values("start_position"),
            get_unique_values("contract_number"), get_unique_values("contract_date"),
            get_unique_values("order_number"), get_unique_values("order_date"),
            employee.bithday, employee.age, employee.gender, employee.email, employee.phone,
            employee.permanent_registration, employee.organization, employee.nationality,
            employee.place_of_birth, employee.city, employee.graduation_end_year, employee.end_education
        ])

    # Javob qaytarish
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=employees_filtered.xlsx'
    workbook.save(response)
    return response

@login_decorator
@role_required("view_reports")
def all_employees(request): 
    employees = Employee.objects.filter(xodim__isnull=False).distinct().prefetch_related("xodim").order_by('-id')  # Barcha xodimlarni olish

    query = request.GET.get('q', '')  # Qidiruv uchun so'rov
    katta_yosh = request.GET.get('katta', None)  # Katta yosh filtri
    kichik_yosh = request.GET.get('kichik', None)  # Kichik yosh filtri
    gender = request.GET.get('gender', None)  # Jinsi
    ilmiy_daraja = request.GET.get('ilmiy_daraja', None)  # Ilmiy daraja
    ilmiy_unvon = request.GET.get('ilmiy_unvon', None)  # Ilmiy unvon
    kafedra = request.GET.get('kafedra', None)  # kafedra
    lavozim = request.GET.get('lavozim', None)  # kafedra
    tugilgan_joyi = request.GET.get('placeofbirtht', None)  # viloyati
    citys = request.GET.get('city', None)  # Doimiy yashash joyi

    academic_degree = Employee.objects.filter(academic_degree__isnull=False).values_list('academic_degree', flat=True).distinct() # Ilmiy darajalar
    academic_title = Employee.objects.filter(academic_title__isnull=False).values_list('academic_title', flat=True).distinct() # Ilmiy unvonlar
    department = Departments.objects.filter(department__isnull=False).values_list('department', flat=True).distinct() # Kafedralar
    labor_form = Departments.objects.filter(labor_form__isnull=False).values_list('labor_form', flat=True).distinct() # Kafedralar
    place_of_birth = Employee.objects.filter(place_of_birth__isnull=False).values_list('place_of_birth', flat=True).distinct() # Tug'ilgan joylar
    city = Employee.objects.filter(city__isnull=False).values_list('city', flat=True).distinct() # Doimiy yashash joyi
    
    # Filtrlash
    if query:
        employees = employees.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(sur_name__icontains=query) |
            Q(employee_id__icontains=query) |
            Q(phone__icontains=query) |
            Q(age__icontains=query) 
        )
    if katta_yosh:
        employees = employees.filter(age__gte=int(katta_yosh))  # Yoshdan katta
    if kichik_yosh:
        employees = employees.filter(age__lte=int(kichik_yosh))  # Yoshdan kichik
    if gender:
        employees = employees.filter(gender=gender)  # Jinsi bo'yicha filtr
    if ilmiy_daraja:
        employees = employees.filter(academic_degree__icontains=ilmiy_daraja)  # Ilmiy daraja bo'yicha filtr
    if ilmiy_unvon:
        employees = employees.filter(academic_title__icontains=ilmiy_unvon)  # Ilmiy unvon bo'yicha filtr
    if kafedra:
        employees = employees.filter(xodim__department__icontains=kafedra)  # Kafedra bo'yicha filtr
    if lavozim:
        employees = employees.filter(xodim__labor_form__icontains=lavozim)  # lavozim bo'yicha filtr
    if tugilgan_joyi:
        employees = employees.filter(place_of_birth__icontains=tugilgan_joyi)  # Tug'ilgan joyi filtr
    if citys:
        employees = employees.filter(city__icontains=city) # Doimiy yashash joyi filtr

    employee_data = []
    for employee in employees:
        departments = Departments.objects.filter(employees=employee)

        employee_data.append({
            "employee_id": employee.employee_id,
            "fio": f"{employee.last_name} {employee.first_name}",
            "sur_name":f"{employee.sur_name}",
            "place_of_birth": employee.place_of_birth,
            "academic_degree_title": f"{employee.academic_degree}, {employee.academic_title}",
            "position": ", ".join(set(departments.values_list("position", flat=True))),
            "labor_form": ", ".join(set(departments.values_list("labor_form", flat=True))),
            "department": ", ".join(set(departments.values_list("department", flat=True))),
            "age": employee.age,
            "gender": employee.gender
        })

    # Pagination
    page_size = request.GET.get('page_size', 10)  # Default 10
    paginator = Paginator(employee_data, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'segment': 'employee',
        "academic_degree": academic_degree,
        "academic_title": academic_title,
        "department": department,  # Kafedralar
        "labor_form":labor_form,
        "place_of_birth": place_of_birth,  # Tug'ilgan joylar
        "city": city,  
    }
    return render(request, 'employee/all_employees.html', context=context)


@login_decorator
@role_required("view_reports")
def view_employees(request, employee_id):  # Xodim ma'lumotlarini ko'rish
    try:
        # Xodimni olish
        employee = Employee.objects.filter(employee_id=employee_id).first()
        user_data = get_employment_info(employee_id)
        user_info = user_data.get('data', {}).get('items', [])[0]

        departments = Departments.objects.filter(employees=employee_id)
        
        # Xodimga bog‘langan departments ma'lumotlarini olish va dublikatlarni yo'qotish
        employee_data = []
        departments = Departments.objects.filter(employees=employee)
        
        # None qiymatlariga tekshirish qo'shish
        employee_data.append({
            "faculty": ", ".join(filter(None, set(departments.values_list("faculty", flat=True)))),
            "department": ", ".join(filter(None, set(departments.values_list("department", flat=True)))),
            "labor_form": ", ".join(filter(None, set(departments.values_list("labor_form", flat=True)))),
            "stavka": ", ".join(filter(None, set(departments.values_list("stavka", flat=True)))),
            "position": ", ".join(filter(None, set(departments.values_list("position", flat=True)))),
            "expertise": ", ".join(filter(None, set(departments.values_list("expertise", flat=True)))),
            "start_position": ", ".join(filter(None, set(departments.values_list("start_position", flat=True)))),
            "contract_number": ", ".join(filter(None, set(departments.values_list("contract_number", flat=True)))),
            "contract_date": ", ".join(filter(None, set(departments.values_list("contract_date", flat=True)))),
            "order_number": ", ".join(filter(None, set(departments.values_list("order_number", flat=True)))),
            "order_date": ", ".join(filter(None, set(departments.values_list("order_date", flat=True)))),
        })
    
    except Employee.DoesNotExist:
        return HttpResponseNotFound("Xodim topilmadi.")  # 404 xatolikni chiqarish

    context = {
        "employee": employee,
        "departments": departments,
        "segment": "employee",
        "user_info": user_info,
        "employee_data": employee_data,  # Xodim bog‘langan unikal departments
    }

    return render(request, "employee/view_employees.html", context=context)

 # Xodim ma'lumotlarini chiqarish

@login_decorator
@role_required("add_user")
def add_employees(request): # Xodim qo'shish
    if request.method == "POST": # POST so'rovni tekshirish
        employee_id = request.POST.get("employee_id")
        citizenship = request.POST.get("citizenship")
        passport = request.POST.get("passport")
        personal_number = request.POST.get("personal_number")
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        sur_name = request.POST.get("sur_name")
        academic_degree = request.POST.get("academic_degree")
        academic_title = request.POST.get("academic_title")
        bithday = request.POST.get("bithday")
        gender = request.POST.get("gender")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        permanent_registration = request.POST.get("permanent_registration")
        organization = request.POST.get("organization")
        nationality = request.POST.get("nationality")
        place_of_birth = request.POST.get("place_of_birth")
        city = request.POST.get("city")
        graduation_end_year = request.POST.get("graduation_end_year")
        end_education = request.POST.get("end_education")
        
        employee = Employee.objects.create(
            employee_id=employee_id, citizenship=citizenship, passport=passport, personal_number=personal_number,
            first_name=first_name, last_name=last_name, sur_name=sur_name, academic_degree=academic_degree,
            academic_title=academic_title, bithday=bithday, gender=gender,
            email=email, phone=phone, permanent_registration=permanent_registration,
            organization=organization, nationality=nationality, place_of_birth=place_of_birth,
            city=city, graduation_end_year=graduation_end_year, end_education=end_education,
        ) # Xodimni yaratish

        return redirect("all_employees")  # Talabalar ro'yxatiga yo'naltirish
    context = {'segment':'employee'} 
    return render(request, "employee/add_employees.html", context=context) # Xodim qo'shish sahifasini chiqarish

@login_decorator
@role_required("ubdate_user")
def edit_employees(request, employee_id): # Xodim ma'lumotlarini tahrirlash
    employee = Employee.objects.filter(employee_id=employee_id).first() # Xodimni olish

    if request.method == "POST": 
        # Talabaning asosiy ma’lumotlarini yangilash
        employee.employee_id = request.POST.get("employee_id")
        employee.citizenship = request.POST.get("citizenship")
        employee.passport = request.POST.get("passport")
        employee.personal_number = request.POST.get("personal_number")
        employee.first_name = request.POST.get("first_name")
        employee.last_name = request.POST.get("last_name")
        employee.sur_name = request.POST.get("sur_name")
        employee.academic_degree = request.POST.get("academic_degree")
        employee.academic_title = request.POST.get("academic_title")
        employee.bithday = request.POST.get("bithday")
        employee.gender = request.POST.get("gender")
        employee.email = request.POST.get("email")
        employee.phone = request.POST.get("phone")
        employee.permanent_registration = request.POST.get("permanent_registration")
        employee.organization = request.POST.get("organization")
        employee.nationality = request.POST.get("nationality")
        employee.place_of_birth = request.POST.get("place_of_birth")
        employee.city = request.POST.get("city")
        employee.graduation_end_year = request.POST.get("graduation_end_year")
        employee.end_education = request.POST.get("end_education")
        employee.save() # Ma'lumotlarni saqlash


        return redirect("all_employees")  # Talabalar ro'yxatiga yo'naltirish
    context = {
        "employee": employee, 'segment':'employee'
        }

    return render(request, "employee/add_employees.html", context=context) # Xodim ma'lumotlarini tahrirlash sahifasini chiqarish

@login_decorator
@role_required("ubdate_user")
def add_department(request, employee_id):
    # Xodimlar ro‘yxatini olish
    employees = Employee.objects.all()

    # Agar employee_id berilgan bo‘lsa, tanlangan xodimni olish
    if employee_id:
        selected_employee = get_object_or_404(Employee, employee_id=employee_id)
    else:
        selected_employee = None

    if request.method == "POST":
        employee = Employee.objects.get(id=request.POST.get("employees"))
        faculty = request.POST.get("faculty")
        department_name = request.POST.get("department")
        labor_form = request.POST.get("labor_form")
        stavka = request.POST.get("stavka")
        position = request.POST.get("position")
        expertise = request.POST.get("expertise")
        start_position = request.POST.get("start_position")
        contract_number = request.POST.get("contract_number")
        contract_date = request.POST.get("contract_date")
        order_number = request.POST.get("order_number")
        order_date = request.POST.get("order_date")

        # Yangi bo'limni yaratish
        Departments.objects.create(
            employees=employee,
            faculty=faculty,
            department=department_name,
            labor_form=labor_form,
            stavka=stavka,
            position=position,
            expertise=expertise,
            start_position=start_position,
            contract_number=contract_number,
            contract_date=contract_date,
            order_number=order_number,
            order_date=order_date
        )

        # Bo‘limlar ro‘yxatiga qaytish
        return redirect("all_employees")

    # Formani yaratish uchun kontekst
    context = {
        "employees": employees,
        "selected_employee": selected_employee,  # Tanlangan xodimni uzatish
        "segment": "department",
        "employee_id":employee_id
    }

    return render(request, "employee/add_department.html", context=context)


@login_decorator
@role_required("ubdate_user")
def edit_departments(request, department_id, employee_id):
    department = Departments.objects.filter(id=department_id).first()
    employees = Employee.objects.all()  # Xodimlarni olish

    if employee_id:
        selected_employee = get_object_or_404(Employee, employee_id=employee_id)
    else:
        selected_employee = None
    
    if request.method == "POST":
        employees_id = request.POST.get("employees")
        faculty = request.POST.get("faculty")
        department_name = request.POST.get("department")

        if not employees_id or not faculty or not department_name:
            messages.error(request, "Barcha maydonlarni to‘ldirish majburiy!")
            return redirect("edit_department", department_id=department.id)

        department.employees = Employee.objects.filter(id=employees_id).first()
        department.faculty = faculty
        department.department = department_name
        department.labor_form = request.POST.get("labor_form", "")
        department.stavka = request.POST.get("stavka", "")
        department.position = request.POST.get("position", "")
        department.expertise = request.POST.get("expertise", "")
        department.start_position = request.POST.get("start_position", "")
        department.contract_number = request.POST.get("contract_number", "")
        department.contract_date = request.POST.get("contract_date", None)
        department.order_number = request.POST.get("order_number", "")
        department.order_date = request.POST.get("order_date", None)

        department.save()

        messages.success(request, "Bo‘lim ma'lumotlari muvaffaqiyatli yangilandi.")
        return redirect("all_employees")

    context = {
        "department": department,
        "employees": employees,  # Xodimlar ro‘yxatini context ga qo‘shish
        'segment': 'department', 
        "selected_employee": selected_employee,
        "employee_id":employee_id
    }

    return render(request, "employee/add_department.html", context=context)


@login_decorator
@role_required("delete_user")
def delete_departments(request, employee_id): # Xodimni o'chirish
    employee = Departments.objects.filter(id=employee_id).first() # Xodimni olish
    employee.delete() # Xodimni o'chirish
    return redirect("all_employees") # Talabalar ro'yxatiga yo'naltirish


@login_decorator
@role_required("delete_user")
def delete_employees(request, employee_id): # Xodimni o'chirish
    employee = Employee.objects.filter(employee_id=employee_id).first() # Xodimni olish
    employee.delete() # Xodimni o'chirish
    return redirect("all_employees") # Talabalar ro'yxatiga yo'naltirish

@login_decorator
def manage_roles(request):
    if request.method == 'POST':
        role_name = request.POST['role_name']
        permissions = request.POST['permissions']  # JSON formatda
        Role.objects.create(name=role_name, permissions=permissions)
        return redirect('manage_roles')

    roles = Role.objects.all()
    return render(request, 'admin/manage_roles.html', {'roles': roles})

@login_decorator
def assign_role(request, user_id):
    user = User.objects.get(id=user_id)
    if request.method == 'POST':
        role_id = request.POST['role']
        role = Role.objects.get(id=role_id)
        UserRole.objects.update_or_create(user=user, defaults={'role': role})
        return redirect('assign_role', user_id=user.id)

    roles = Role.objects.all()
    return render(request, 'admin/assign_role.html', {'user': user, 'roles': roles})
